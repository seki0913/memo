import openpyxl as px

wb = px.Workbook()

ws = wb["Sheet"]
ws = wb.remove_sheet(ws)
ws = wb.create_sheet(title="パラメータシート") 

#ws = wb.active

# シート名を指定

# セル基本操作
ws["A1"].value = "Hello World"
ws["A2"].value = "Hello World"


wb.save("./test.xlsx")